from fastapi import FastAPI, Request, UploadFile, File
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, validator
from jose import jwt, JWTError
from passlib.context import CryptContext
import psycopg2, os, time, io, csv, re
from dotenv import load_dotenv
from collections import defaultdict
from threading import Lock

load_dotenv()
app = FastAPI()
SECRET_KEY = os.environ.get("SECRET_KEY","fallback-secret")
ALGORITHM = "HS256"
TOKEN_EXP = 86400 * 3  # 3 dias (antes era 7)
pwd_ctx = CryptContext(schemes=["bcrypt"],deprecated="auto")
ALLOWED_ORIGINS = os.environ.get("ALLOWED_ORIGINS","https://errtrack-526z.onrender.com").split(",")
app.add_middleware(CORSMiddleware,allow_origins=ALLOWED_ORIGINS,allow_credentials=True,allow_methods=["*"],allow_headers=["*"])
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app.mount("/js",StaticFiles(directory=os.path.join(BASE_DIR,"front-end/js")),name="js")
app.mount("/css",StaticFiles(directory=os.path.join(BASE_DIR,"front-end/css")),name="css")
app.mount("/img",StaticFiles(directory=os.path.join(BASE_DIR,"front-end/img")),name="img")

# ── RATE LIMITING ─────────────────────────────────────────────────────────────
_login_attempts = defaultdict(list)  # ip -> [timestamps]
_login_lock = Lock()
_token_blacklist = set()  # tokens invalidados no logout

MAX_ATTEMPTS = 5    # máximo de tentativas
WINDOW_SEC   = 300  # janela de 5 minutos
BLOCK_SEC    = 900  # bloqueio de 15 minutos

def get_client_ip(request: Request) -> str:
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"

def check_rate_limit(ip: str) -> tuple[bool, int]:
    """Retorna (bloqueado, segundos_restantes)"""
    now = time.time()
    with _login_lock:
        attempts = _login_attempts[ip]
        # Remove tentativas fora da janela
        attempts = [t for t in attempts if now - t < WINDOW_SEC]
        _login_attempts[ip] = attempts
        if len(attempts) >= MAX_ATTEMPTS:
            oldest = attempts[0]
            remaining = int(BLOCK_SEC - (now - oldest))
            if remaining > 0:
                return True, remaining
            else:
                _login_attempts[ip] = []
                return False, 0
        return False, 0

def register_attempt(ip: str):
    with _login_lock:
        _login_attempts[ip].append(time.time())

def clear_attempts(ip: str):
    with _login_lock:
        _login_attempts[ip] = []

# ── VALIDAÇÃO DE SENHA ────────────────────────────────────────────────────────
def validar_senha(senha: str) -> tuple[bool, str]:
    if len(senha) < 8:
        return False, "Senha deve ter pelo menos 8 caracteres."
    if not re.search(r'[A-Z]', senha):
        return False, "Senha deve conter pelo menos uma letra maiúscula."
    if not re.search(r'[0-9]', senha):
        return False, "Senha deve conter pelo menos um número."
    if not re.search(r'[^A-Za-z0-9]', senha):
        return False, "Senha deve conter pelo menos um caractere especial."
    return True, ""

# ── DB ────────────────────────────────────────────────────────────────────────
_conn = None
def get_conn():
    global _conn
    try:
        if _conn is None or _conn.closed: raise Exception()
        _conn.cursor().execute("SELECT 1")
    except:
        _conn = psycopg2.connect(os.environ["DATABASE_URL"])
        _conn.autocommit = False
    return _conn
def get_cursor(): return get_conn().cursor()
def commit(): get_conn().commit()

def _criatabela():
    cur = get_cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS admins(id SERIAL PRIMARY KEY,usuario TEXT UNIQUE,senha TEXT,role TEXT DEFAULT 'admin')")
    cur.execute("CREATE TABLE IF NOT EXISTS funcionarios(id SERIAL PRIMARY KEY,nomecompleto TEXT,especializacao TEXT,periodotrabalho TEXT,categoria TEXT,observacoes TEXT,pausa1 TEXT,pausa2 TEXT,pausa3 TEXT)")
    cur.execute("CREATE TABLE IF NOT EXISTS erros(id SERIAL PRIMARY KEY,nomefuncionario TEXT,periodo TEXT,descricao TEXT,gravidade TEXT,categoria TEXT,ts INTEGER)")
    cur.execute("CREATE TABLE IF NOT EXISTS feedbacks(id SERIAL PRIMARY KEY,nomefuncionario TEXT,nota_geral REAL,pontos_melhora TEXT,texto_feedback TEXT,aplicado_por TEXT,ts INTEGER)")
    cur.execute("CREATE TABLE IF NOT EXISTS treinamentos(id SERIAL PRIMARY KEY,nomefuncionario TEXT,titulo TEXT,descricao TEXT,status TEXT DEFAULT 'pendente',aplicado_por TEXT,ts INTEGER)")
    cur.execute("CREATE TABLE IF NOT EXISTS indicadores(id SERIAL PRIMARY KEY,fila TEXT,segmento TEXT,tma TEXT,tme TEXT,total INTEGER,atendidas INTEGER,perdidas INTEGER,sla TEXT,periodo TEXT,ts INTEGER)")
    cur.execute("""CREATE TABLE IF NOT EXISTS logs_acesso(
        id SERIAL PRIMARY KEY,
        usuario TEXT,
        acao TEXT,
        detalhe TEXT,
        ip TEXT,
        ts INTEGER
    )""")
    for col in ["pausa1","pausa2","pausa3","observacoes","especializacao","periodotrabalho"]:
        try: cur.execute(f"ALTER TABLE funcionarios ADD COLUMN IF NOT EXISTS {col} TEXT DEFAULT ''")
        except: get_conn().rollback()
    commit()

@app.on_event("startup")
def startup(): _criatabela()

# ── LOGS DE ACESSO ────────────────────────────────────────────────────────────
def log_acao(usuario: str, acao: str, detalhe: str = "", ip: str = ""):
    try:
        cur = get_cursor()
        cur.execute("INSERT INTO logs_acesso(usuario,acao,detalhe,ip,ts) VALUES(%s,%s,%s,%s,%s)",
                    (usuario, acao, detalhe[:500], ip, int(time.time())))
        commit()
    except: pass

# ── MODELOS ───────────────────────────────────────────────────────────────────
class LoginData(BaseModel): usuario:str; senha:str
class ErroData(BaseModel): nomefuncionario:str; periodo:str; descricao:str; gravidade:str; categoria:str=""
class FuncionarioData(BaseModel): nomecompleto:str; especializacao:str=""; periodotrabalho:str=""; categoria:str=""; observacoes:str=""; pausa1:str=""; pausa2:str=""; pausa3:str=""
class AdminData(BaseModel): usuario:str; senha:str; pode_criar_admins:bool=False
class Feedback(BaseModel): nomefuncionario:str; nota_geral:float; pontos_melhora:str; texto_feedback:str; aplicado_por:str
class Treinamento(BaseModel): nomefuncionario:str; titulo:str; descricao:str=""; status:str="pendente"; aplicado_por:str

# ── AUTH ──────────────────────────────────────────────────────────────────────
def criar_token(u, r): return jwt.encode({"sub":u,"role":r,"exp":time.time()+TOKEN_EXP,"iat":time.time()},SECRET_KEY,algorithm=ALGORITHM)

def usuario_autenticado(request: Request):
    token = request.cookies.get("token")
    if not token: return None
    if token in _token_blacklist: return None
    try: return jwt.decode(token,SECRET_KEY,algorithms=[ALGORITHM])
    except: return None

# ── HEADERS DE SEGURANÇA ──────────────────────────────────────────────────────
@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    return response

# ── ROTAS PÚBLICAS ────────────────────────────────────────────────────────────
@app.get("/")
def serve_login(): return FileResponse(os.path.join(BASE_DIR,"front-end/login.html"))

@app.get("/sistema")
def serve_sistema(request:Request):
    if not usuario_autenticado(request): return FileResponse(os.path.join(BASE_DIR,"front-end/login.html"))
    path = os.path.join(BASE_DIR,"front-end/errtrack-premium.html")
    with open(path,"r",encoding="utf-8") as f: content = f.read().replace("/static/conectaapi.js","/js/conectaapi.js")
    return HTMLResponse(content=content)

@app.post("/login")
def login(data: LoginData, request: Request):
    ip = get_client_ip(request)

    # Rate limiting
    bloqueado, restante = check_rate_limit(ip)
    if bloqueado:
        log_acao(data.usuario, "LOGIN_BLOQUEADO", f"IP bloqueado por {restante}s", ip)
        return JSONResponse({"status":"erro","mensagem":f"Muitas tentativas. Aguarde {restante} segundos."},status_code=429)

    # Sanitiza input
    usuario = data.usuario.strip()[:50]

    cur = get_cursor()
    cur.execute("SELECT senha,role FROM admins WHERE usuario=%s",(usuario,))
    row = cur.fetchone()

    if not row or not pwd_ctx.verify(data.senha, row[0]):
        register_attempt(ip)
        tentativas_restantes = MAX_ATTEMPTS - len(_login_attempts[ip])
        log_acao(usuario, "LOGIN_FALHOU", f"IP: {ip}", ip)
        return JSONResponse({"status":"erro","mensagem":f"Usuário ou senha incorretos. {tentativas_restantes} tentativa(s) restante(s)."},status_code=401)

    # Login bem-sucedido
    clear_attempts(ip)
    token = criar_token(usuario, row[1])
    log_acao(usuario, "LOGIN_OK", f"IP: {ip}", ip)
    resp = JSONResponse({"status":"sucesso","role":row[1]})
    resp.set_cookie("token",token,httponly=True,samesite="lax",secure=os.environ.get("RENDER")=="true",max_age=TOKEN_EXP)
    return resp

@app.post("/logout")
def logout(request: Request):
    token = request.cookies.get("token")
    if token:
        _token_blacklist.add(token)
        p = usuario_autenticado(request)
        if p: log_acao(p.get("sub","?"), "LOGOUT", "", get_client_ip(request))
    resp = JSONResponse({"status":"sucesso"})
    resp.delete_cookie("token")
    return resp

@app.get("/me")
def me(request:Request):
    p = usuario_autenticado(request)
    if not p: return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    return {"usuario":p.get("sub"),"role":p.get("role")}

@app.post("/setup")
def setup():
    cur=get_cursor(); cur.execute("SELECT COUNT(*) FROM admins")
    if cur.fetchone()[0]>0: return JSONResponse({"status":"erro","mensagem":"Setup já realizado."},status_code=403)
    cur.execute("INSERT INTO admins(usuario,senha,role) VALUES(%s,%s,%s)",("Lucas.Martins",pwd_ctx.hash("Master@2026!"),"superadmin")); commit()
    return {"status":"sucesso","mensagem":"Superadmin criado!"}

@app.post("/reset-senha-temp")
def reset_senha():
    cur=get_cursor(); cur.execute("UPDATE admins SET senha=%s WHERE usuario='Lucas.Martins'",(pwd_ctx.hash("Master@2026!"),)); commit()
    return {"status":"sucesso"} if cur.rowcount else JSONResponse({"mensagem":"Não encontrado."},status_code=404)

@app.post("/migrar")
def migrar(request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor()
    for col in ["pausa1","pausa2","pausa3","observacoes","especializacao","periodotrabalho"]:
        try: cur.execute(f"ALTER TABLE funcionarios ADD COLUMN IF NOT EXISTS {col} TEXT DEFAULT ''")
        except: get_conn().rollback()
    commit(); return {"status":"sucesso","mensagem":"Migração concluída!"}

# ── LOGS ──────────────────────────────────────────────────────────────────────
@app.get("/logs")
def ver_logs(request:Request):
    p = usuario_autenticado(request)
    if not p or p.get("role") not in ("superadmin","admin_full"):
        return JSONResponse({"mensagem":"Sem permissão."},status_code=403)
    cur=get_cursor()
    cur.execute("SELECT usuario,acao,detalhe,ip,ts FROM logs_acesso ORDER BY ts DESC LIMIT 200")
    rows=cur.fetchall()
    return {"logs":[{"usuario":r[0],"acao":r[1],"detalhe":r[2],"ip":r[3],"ts":r[4]} for r in rows]}

# ── ADMINS ────────────────────────────────────────────────────────────────────
@app.get("/admins")
def listar_admins(request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("SELECT usuario,role FROM admins ORDER BY id"); rows=cur.fetchall()
    return {"admins":[{"usuario":r[0],"role":r[1]} for r in rows]}

@app.post("/admins")
def criar_admin(data:AdminData,request:Request):
    p=usuario_autenticado(request)
    if not p or p.get("role") not in ("superadmin","admin_full"): return JSONResponse({"mensagem":"Sem permissão."},status_code=403)
    # Valida senha
    ok, msg = validar_senha(data.senha)
    if not ok: return JSONResponse({"mensagem":msg},status_code=400)
    cur=get_cursor()
    try:
        cur.execute("INSERT INTO admins(usuario,senha,role) VALUES(%s,%s,%s)",(data.usuario.strip(),pwd_ctx.hash(data.senha),"admin_full" if data.pode_criar_admins else "admin"))
        commit()
        log_acao(p.get("sub"), "ADMIN_CRIADO", f"Novo admin: {data.usuario}", get_client_ip(request))
        return {"status":"sucesso","mensagem":f"Admin '{data.usuario}' criado!"}
    except: get_conn().rollback(); return JSONResponse({"mensagem":"Usuário já existe."},status_code=409)

@app.delete("/admins/{usuario}")
def deletar_admin(usuario:str,request:Request):
    p=usuario_autenticado(request)
    if not p or p.get("role") not in ("superadmin","admin_full"): return JSONResponse({"mensagem":"Sem permissão."},status_code=403)
    cur=get_cursor(); cur.execute("SELECT role FROM admins WHERE usuario=%s",(usuario,)); row=cur.fetchone()
    if not row: return JSONResponse({"mensagem":"Não encontrado."},status_code=404)
    if row[0]=="superadmin": return JSONResponse({"mensagem":"Não pode remover superadmin."},status_code=403)
    cur.execute("DELETE FROM admins WHERE usuario=%s",(usuario,)); commit()
    log_acao(p.get("sub"), "ADMIN_REMOVIDO", f"Admin removido: {usuario}", get_client_ip(request))
    return {"status":"sucesso","mensagem":f"Admin '{usuario}' removido!"}

# ── FUNCIONÁRIOS ──────────────────────────────────────────────────────────────
@app.get("/funcionarios")
def listar_funcionarios(request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("SELECT id,nomecompleto,especializacao,periodotrabalho,categoria,observacoes,pausa1,pausa2,pausa3 FROM funcionarios ORDER BY nomecompleto"); rows=cur.fetchall()
    return {"funcionarios":[{"id":r[0],"nomecompleto":r[1],"especializacao":r[2],"periodotrabalho":r[3],"categoria":r[4],"observacoes":r[5],"pausa1":r[6],"pausa2":r[7],"pausa3":r[8]} for r in rows]}

@app.post("/funcionarios")
def criar_funcionario(data:FuncionarioData,request:Request):
    p=usuario_autenticado(request)
    if not p: return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("INSERT INTO funcionarios(nomecompleto,especializacao,periodotrabalho,categoria,observacoes,pausa1,pausa2,pausa3) VALUES(%s,%s,%s,%s,%s,%s,%s,%s)",(data.nomecompleto,data.especializacao,data.periodotrabalho,data.categoria,data.observacoes,data.pausa1,data.pausa2,data.pausa3)); commit()
    log_acao(p.get("sub"), "FUNC_CRIADO", data.nomecompleto, get_client_ip(request))
    return {"status":"sucesso","mensagem":"Funcionário cadastrado!"}

@app.put("/funcionarios/{fid}")
def atualizar_funcionario(fid:int,data:FuncionarioData,request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("UPDATE funcionarios SET nomecompleto=%s,especializacao=%s,periodotrabalho=%s,categoria=%s,observacoes=%s,pausa1=%s,pausa2=%s,pausa3=%s WHERE id=%s",(data.nomecompleto,data.especializacao,data.periodotrabalho,data.categoria,data.observacoes,data.pausa1,data.pausa2,data.pausa3,fid)); commit()
    return {"status":"sucesso","mensagem":"Funcionário atualizado!"}

@app.delete("/funcionarios/{fid}")
def deletar_funcionario(fid:int,request:Request):
    p=usuario_autenticado(request)
    if not p: return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("DELETE FROM funcionarios WHERE id=%s",(fid,)); commit()
    log_acao(p.get("sub"), "FUNC_DELETADO", f"ID: {fid}", get_client_ip(request))
    return {"status":"sucesso","mensagem":"Funcionário removido!"}

@app.post("/importar-pausas")
async def importar_pausas(request:Request,file:UploadFile=File(...)):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    import openpyxl
    wb=openpyxl.load_workbook(io.BytesIO(await file.read())); ws=wb.active; cur=get_cursor(); n=0
    for row in ws.iter_rows(min_row=2,values_only=True):
        if not row[0]: continue
        cur.execute("UPDATE funcionarios SET pausa1=%s,pausa2=%s,pausa3=%s WHERE nomecompleto ILIKE %s",(str(row[1] or ""),str(row[2] or ""),str(row[3] or ""),str(row[0]).strip())); n+=cur.rowcount
    commit(); return {"status":"sucesso","mensagem":f"{n} funcionário(s) atualizado(s)!"}

@app.post("/importar-funcionarios")
async def importar_funcionarios(request:Request,file:UploadFile=File(...)):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    import openpyxl
    wb=openpyxl.load_workbook(io.BytesIO(await file.read())); ws=wb["Funcionários"]; cur=get_cursor(); n=0
    for row in ws.iter_rows(min_row=2,values_only=True):
        if not row[0]: continue
        nome=str(row[0]).strip(); cat=str(row[1] or "").strip(); espec=str(row[2] or "").strip(); per=str(row[3] or "").strip(); obs=str(row[4] or "").strip()
        cur.execute("SELECT id FROM funcionarios WHERE nomecompleto ILIKE %s",(nome,))
        if cur.fetchone(): continue
        cur.execute("INSERT INTO funcionarios(nomecompleto,especializacao,periodotrabalho,categoria,observacoes,pausa1,pausa2,pausa3) VALUES(%s,%s,%s,%s,%s,'','','')",(nome,espec,per,cat,obs)); n+=1
    commit(); return {"status":"sucesso","mensagem":f"{n} funcionários importados!"}

@app.post("/importar-erros")
async def importar_erros_bulk(request:Request,file:UploadFile=File(...)):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    import openpyxl
    wb=openpyxl.load_workbook(io.BytesIO(await file.read())); ws=wb["Erros"]; cur=get_cursor(); n=0
    for row in ws.iter_rows(min_row=2,values_only=True):
        if not row[0]: continue
        cur.execute("INSERT INTO erros(nomefuncionario,periodo,descricao,gravidade,categoria,ts) VALUES(%s,%s,%s,%s,%s,%s)",(str(row[0]).strip(),str(row[4] or ""),str(row[2] or ""),str(row[1] or "media"),str(row[3] or ""),int(time.time()))); n+=1
    commit(); return {"status":"sucesso","mensagem":f"{n} erros importados!"}

# ── ERROS ─────────────────────────────────────────────────────────────────────
@app.get("/erros")
def listar_erros(request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("SELECT e.id,e.nomefuncionario,e.periodo,e.descricao,e.gravidade,e.categoria,e.ts,f.categoria FROM erros e LEFT JOIN funcionarios f ON f.nomecompleto=e.nomefuncionario ORDER BY e.ts DESC"); rows=cur.fetchall()
    return {"erros":[{"id":r[0],"nomefuncionario":r[1],"periodo":r[2],"descricao":r[3],"gravidade":r[4],"categoria":r[5],"ts":r[6],"cat_func":r[7]} for r in rows]}

@app.get("/erros/{nome}")
def erros_por_funcionario(nome:str,request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("SELECT e.id,e.nomefuncionario,e.periodo,e.descricao,e.gravidade,e.categoria,e.ts,f.categoria FROM erros e LEFT JOIN funcionarios f ON f.nomecompleto=e.nomefuncionario WHERE e.nomefuncionario=%s ORDER BY e.ts DESC",(nome,)); rows=cur.fetchall()
    return {"erros":[{"id":r[0],"nomefuncionario":r[1],"periodo":r[2],"descricao":r[3],"gravidade":r[4],"categoria":r[5],"ts":r[6],"cat_func":r[7]} for r in rows]}

@app.post("/erros")
def registrar_erro(data:ErroData,request:Request):
    p=usuario_autenticado(request)
    if not p: return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    # Sanitiza descrição
    desc = data.descricao.strip()[:1000]
    cur=get_cursor(); cur.execute("INSERT INTO erros(nomefuncionario,periodo,descricao,gravidade,categoria,ts) VALUES(%s,%s,%s,%s,%s,%s)",(data.nomefuncionario,data.periodo,desc,data.gravidade,data.categoria,int(time.time()))); commit()
    log_acao(p.get("sub"), "ERRO_REGISTRADO", f"{data.nomefuncionario} · {data.gravidade}", get_client_ip(request))
    return {"status":"sucesso","mensagem":"Erro registrado!"}

@app.delete("/erros/{erro_id}")
def deletar_erro(erro_id:int,request:Request):
    p=usuario_autenticado(request)
    if not p: return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("DELETE FROM erros WHERE id=%s",(erro_id,)); commit()
    log_acao(p.get("sub"), "ERRO_DELETADO", f"ID: {erro_id}", get_client_ip(request))
    return {"status":"sucesso"} if cur.rowcount else JSONResponse({"mensagem":"Não encontrado."},status_code=404)

# ── EXPORTAR ──────────────────────────────────────────────────────────────────
@app.get("/exportar-excel")
def exportar_excel(request:Request):
    p=usuario_autenticado(request)
    if not p: return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("SELECT nomefuncionario,periodo,descricao,gravidade,categoria,to_timestamp(ts) FROM erros ORDER BY ts DESC"); rows=cur.fetchall()
    out=io.StringIO(); out.write('\ufeff'); out.write('Funcionário,Período,Descrição,Gravidade,Categoria,Data\n')
    for r in rows: out.write(','.join([f'"{str(x or "").replace(chr(34),chr(39))}"' for x in r])+'\n')
    out.seek(0)
    log_acao(p.get("sub"), "EXPORTOU", "CSV de erros", get_client_ip(request))
    return StreamingResponse(io.BytesIO(out.getvalue().encode('utf-8')),media_type='text/csv',headers={'Content-Disposition':'attachment; filename=errtrack_export.csv'})

# ── FEEDBACKS ─────────────────────────────────────────────────────────────────
@app.get("/feedbacks")
def listar_feedbacks(request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("SELECT id,nomefuncionario,nota_geral,pontos_melhora,texto_feedback,aplicado_por,ts FROM feedbacks ORDER BY ts DESC"); rows=cur.fetchall()
    return {"feedbacks":[{"id":r[0],"nomefuncionario":r[1],"nota_geral":r[2],"pontos_melhora":r[3],"texto_feedback":r[4],"aplicado_por":r[5],"ts":r[6]} for r in rows]}

@app.get("/feedbacks/{nome}")
def feedbacks_operador(nome:str,request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("SELECT id,nomefuncionario,nota_geral,pontos_melhora,texto_feedback,aplicado_por,ts FROM feedbacks WHERE nomefuncionario=%s ORDER BY ts DESC",(nome,)); rows=cur.fetchall()
    return {"feedbacks":[{"id":r[0],"nomefuncionario":r[1],"nota_geral":r[2],"pontos_melhora":r[3],"texto_feedback":r[4],"aplicado_por":r[5],"ts":r[6]} for r in rows]}

@app.post("/feedbacks")
def salvar_feedback(fb:Feedback,request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("INSERT INTO feedbacks(nomefuncionario,nota_geral,pontos_melhora,texto_feedback,aplicado_por,ts) VALUES(%s,%s,%s,%s,%s,%s)",(fb.nomefuncionario,fb.nota_geral,fb.pontos_melhora,fb.texto_feedback,fb.aplicado_por,int(time.time()))); commit()
    return {"status":"sucesso","mensagem":"Feedback salvo!"}

@app.delete("/feedbacks/{fid}")
def deletar_feedback(fid:int,request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("DELETE FROM feedbacks WHERE id=%s",(fid,)); commit()
    return {"status":"sucesso","mensagem":"Feedback removido!"}

# ── TREINAMENTOS ──────────────────────────────────────────────────────────────
@app.get("/treinamentos")
def listar_treinamentos(request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("SELECT id,nomefuncionario,titulo,descricao,status,aplicado_por,ts FROM treinamentos ORDER BY ts DESC"); rows=cur.fetchall()
    return {"treinamentos":[{"id":r[0],"nomefuncionario":r[1],"titulo":r[2],"descricao":r[3],"status":r[4],"aplicado_por":r[5],"ts":r[6]} for r in rows]}

@app.post("/treinamentos")
def criar_treinamento(t:Treinamento,request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("INSERT INTO treinamentos(nomefuncionario,titulo,descricao,status,aplicado_por,ts) VALUES(%s,%s,%s,%s,%s,%s)",(t.nomefuncionario,t.titulo,t.descricao,t.status,t.aplicado_por,int(time.time()))); commit()
    return {"status":"sucesso","mensagem":"Treinamento criado!"}

@app.put("/treinamentos/{tid}")
async def atualizar_treinamento(tid:int,request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    body=await request.json(); cur=get_cursor(); cur.execute("UPDATE treinamentos SET status=%s WHERE id=%s",(body.get("status"),tid)); commit()
    return {"status":"sucesso"}

@app.delete("/treinamentos/{tid}")
def deletar_treinamento(tid:int,request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("DELETE FROM treinamentos WHERE id=%s",(tid,)); commit()
    return {"status":"sucesso","mensagem":"Treinamento removido!"}

# ── INDICADORES ───────────────────────────────────────────────────────────────
FILA_SEGMENTO={"ALTERAR VENCIMENTO DE FATURA":"SAC","DUVIDAS DE COBRANÇA":"SAC","NEGOCIAR DÉBITOS":"SAC","SEGUNDA VIA DE BOLETO":"SAC","LIBERAÇÃO DE CONFIANÇA":"SAC","FALAR COM ATENDENTE":"SAC","CONFIGURAR ROTEADOR":"SUPORTE","LENTIDÃO INTERNET":"SUPORTE","SEM SINAL DE INTERNET":"SUPORTE","IMAGENS RUINS":"PRODUTOS DIGITAIS","PRODUTOS DIGITAIS":"PRODUTOS DIGITAIS","TV SEM SINAL":"PRODUTOS DIGITAIS"}

@app.post("/indicadores/upload")
async def upload_indicadores(request:Request,file:UploadFile=File(...)):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    text=(await file.read()).decode("utf-8-sig",errors="ignore"); reader=csv.DictReader(io.StringIO(text),delimiter=";")
    periodo=time.strftime("%m/%Y"); cur=get_cursor(); cur.execute("DELETE FROM indicadores WHERE periodo=%s",(periodo,)); n=0
    for row in reader:
        fila=(row.get("Fila") or "").strip().upper()
        if not fila: continue
        try: cur.execute("INSERT INTO indicadores(fila,segmento,tma,tme,total,atendidas,perdidas,sla,periodo,ts) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",(fila,FILA_SEGMENTO.get(fila,"OUTROS"),row.get("TMA","").strip(),row.get("TME","").strip(),int(row.get("Total") or 0),int(row.get("Atendidas") or 0),int(row.get("Perdidas") or 0),row.get("%Atendidas <20s","").strip(),periodo,int(time.time()))); n+=1
        except: continue
    commit(); return {"status":"sucesso","mensagem":f"{n} filas importadas para {periodo}!"}

@app.get("/indicadores")
def listar_indicadores(request:Request,periodo:str=None):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor()
    if periodo: cur.execute("SELECT id,fila,segmento,tma,tme,total,atendidas,perdidas,sla,periodo FROM indicadores WHERE periodo=%s ORDER BY segmento,fila",(periodo,))
    else: cur.execute("SELECT id,fila,segmento,tma,tme,total,atendidas,perdidas,sla,periodo FROM indicadores ORDER BY ts DESC LIMIT 100")
    rows=cur.fetchall()
    return {"indicadores":[{"id":r[0],"fila":r[1],"segmento":r[2],"tma":r[3],"tme":r[4],"total":r[5],"atendidas":r[6],"perdidas":r[7],"sla":r[8],"periodo":r[9]} for r in rows]}

@app.get("/indicadores/periodos")
def listar_periodos(request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("SELECT DISTINCT periodo FROM indicadores ORDER BY periodo DESC"); rows=cur.fetchall()
    return {"periodos":[r[0] for r in rows]}

@app.post("/indicadores/limpar")
def limpar_indicadores(request:Request):
    if not usuario_autenticado(request): return JSONResponse({"mensagem":"Não autorizado."},status_code=401)
    cur=get_cursor(); cur.execute("DELETE FROM indicadores WHERE id NOT IN (SELECT MIN(id) FROM indicadores GROUP BY fila,periodo)"); n=cur.rowcount; commit()
    return {"status":"sucesso","mensagem":f"{n} duplicatas removidas!"}
